from __future__ import annotations

import hashlib
import re
import sqlite3
import warnings
import zipfile
import posixpath
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style.*",
    category=UserWarning,
    module="openpyxl.styles.stylesheet",
)
warnings.filterwarnings(
    "ignore",
    message="Conditional Formatting extension is not supported.*",
    category=UserWarning,
    module="openpyxl.worksheet._reader",
)

GENERATED_SHEETS = {"Дашборд", "Сводка", "События"}

EVENT_COLUMNS = [
    "vehicle", "plate", "route", "violation", "action", "severity", "risk", "dt", "start_raw",
    "date", "hour", "start_pos", "start_lat", "start_lon", "start_map_url",
    "end_raw", "end_pos", "end_lat", "end_lon", "end_map_url",
    "value", "value_num", "count"
]

DB_EVENT_COLUMNS = EVENT_COLUMNS + ["event_hash", "source_file", "source_sheet", "file_hash", "imported_at"]


# =========================
# Excel parsing
# =========================

def get_sheet_names(file_bytes: bytes) -> list[str]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    return wb.sheetnames


def pick_source_sheet_name(sheet_names: list[str], preferred: str | None = None) -> str:
    if preferred and preferred in sheet_names:
        return preferred
    if "Качество вождения" in sheet_names:
        return "Качество вождения"
    for name in sheet_names:
        low = name.lower()
        if "качество" in low and "вожд" in low:
            return name
    for name in sheet_names:
        if name not in GENERATED_SHEETS:
            return name
    return sheet_names[0]


def parse_dt(value: Any):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def safe_int(value: Any, default: int = 0) -> int:
    if value in (None, "", "-----"):
        return default
    try:
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return default


def parse_num_value(value: Any):
    if value in (None, "", "-----"):
        return None
    match = re.search(r"[-+]?\d+(?:[\.,]\d+)?", str(value))
    if not match:
        return None
    return float(match.group(0).replace(",", "."))


def split_vehicle(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    plate = text.split()[0] if text else ""
    match = re.search(r"\((.*?)\)", text)
    route = match.group(1).strip() if match else ""
    return plate, route


def action_type(violation: Any) -> str:
    text = str(violation or "").strip()
    low = text.lower()
    if low.startswith("превышение скорости"):
        return "Превышение скорости"
    return text.split(":", 1)[0].strip() if ":" in text else text



def severity_class(violation: Any) -> str:
    text = str(violation or "").lower()

    # Скоростные нарушения Wialon:
    # Превышение скорости (0-10 км/ч): среднее превышение скорости
    # Превышение скорости (10-20 км/ч): опасное превышение скорости
    # Превышение скорости (+20 км/ч): критическое превышение скорости
    if "превышение скорости" in text:
        if "+20" in text or "крит" in text:
            return "Критическое"
        if "10-20" in text or "опас" in text:
            return "Опасное"
        if "0-10" in text or "сред" in text:
            return "Среднее"

    if "крит" in text:
        return "Критическое"
    if "опас" in text:
        return "Опасное"
    if "сред" in text:
        return "Среднее"
    if "резк" in text:
        return "Резкое"
    return "Прочее"



def risk_points(severity: str) -> int:
    weights = {
        "Критическое": 5,
        "Опасное": 3,
        "Среднее": 2,
        "Резкое": 1,
        "Прочее": 0,
    }
    return weights.get(str(severity or ""), 0)



def priority_zone(risk_score: int) -> str:
    if risk_score >= 40:
        return "Красная зона"
    if risk_score >= 15:
        return "Жёлтая зона"
    return "Норма"


def _col_letters_to_index(letters: str) -> int:
    value = 0
    for ch in letters.upper():
        value = value * 26 + (ord(ch) - ord("A") + 1)
    return value - 1


def _cell_ref_to_indexes(ref: str) -> tuple[int, int] | None:
    match = re.match(r"^([A-Z]+)(\d+)$", str(ref or "").upper())
    if not match:
        return None
    col_idx = _col_letters_to_index(match.group(1))
    row_idx = int(match.group(2)) - 1
    return row_idx, col_idx


def _read_hyperlinks(file_bytes: bytes, sheet_name: str) -> dict[tuple[int, int], str]:
    """Extract external hyperlinks for a worksheet directly from XLSX XML.

    This is much faster than loading the workbook in normal openpyxl mode.
    """
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    rel_tag = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"

    links: dict[tuple[int, int], str] = {}
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
            workbook = ET.fromstring(zf.read("xl/workbook.xml"))
            rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

            rid_to_target: dict[str, str] = {}
            for rel in rels.findall(rel_tag):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rid and target:
                    rid_to_target[rid] = target

            sheet_target = None
            for sheet in workbook.findall(f".//{ns_main}sheet"):
                if sheet.attrib.get("name") == sheet_name:
                    rid = sheet.attrib.get(f"{ns_rel}id")
                    target = rid_to_target.get(rid or "")
                    if target:
                        sheet_target = "xl/" + target.lstrip("/")
                    break

            if not sheet_target:
                return links

            sheet_target = posixpath.normpath(sheet_target)
            sheet_dir = posixpath.dirname(sheet_target)
            sheet_file = posixpath.basename(sheet_target)
            rels_path = posixpath.join(sheet_dir, "_rels", sheet_file + ".rels")

            rid_to_url: dict[str, str] = {}
            if rels_path in zf.namelist():
                sheet_rels = ET.fromstring(zf.read(rels_path))
                for rel in sheet_rels.findall(rel_tag):
                    rid = rel.attrib.get("Id")
                    target = rel.attrib.get("Target")
                    mode = rel.attrib.get("TargetMode")
                    if rid and target and mode == "External":
                        rid_to_url[rid] = target

            if not rid_to_url:
                return links

            sheet_xml = ET.fromstring(zf.read(sheet_target))
            for hyperlink in sheet_xml.findall(f".//{ns_main}hyperlink"):
                ref = hyperlink.attrib.get("ref")
                rid = hyperlink.attrib.get(f"{ns_rel}id")
                url = rid_to_url.get(rid or "")
                if not ref or not url:
                    continue
                # Single-cell links are expected in Wialon reports. For ranges, use first cell.
                first_ref = ref.split(":", 1)[0]
                indexes = _cell_ref_to_indexes(first_ref)
                if indexes:
                    links[indexes] = url
    except Exception:
        return {}

    return links


def _read_rows(file_bytes: bytes, sheet_name: str):
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[sheet_name]
    if hasattr(ws, "reset_dimensions"):
        ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    links = _read_hyperlinks(file_bytes, sheet_name)
    return rows, links


def _cell_link(links: dict[tuple[int, int], str], row_idx: int, colmap: dict[str, int], name: str) -> str:
    idx = colmap.get(name)
    if idx is None:
        return ""
    return links.get((row_idx, idx), "")


def _coords_from_map_url(url: Any) -> tuple[float | None, float | None]:
    text = str(url or "")
    match = re.search(r"[?&]q=(-?\d+(?:\.\d+)?),\s*(-?\d+(?:\.\d+)?)", text)
    if not match:
        match = re.search(r"/@(-?\d+(?:\.\d+)?),\s*(-?\d+(?:\.\d+)?)", text)
    if not match:
        return None, None
    try:
        return float(match.group(1)), float(match.group(2))
    except Exception:
        return None, None


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    aliases = {
        "number": {"№", "номер", "n"},
        "vehicle": {"группировка", "тс", "объект", "машина", "unit"},
        "rating": {"рейтинг по нарушениям", "рейтинг"},
        "violation": {"нарушение", "событие", "тип нарушения"},
        "start_raw": {"начало", "дата начала", "время начала", "start"},
        "start_pos": {"нач. положение", "нач положение", "начальное положение", "начальная позиция", "start location"},
        "end_raw": {"конец", "дата конца", "время конца", "end"},
        "end_pos": {"кон. положение", "кон положение", "конечное положение", "конечная позиция", "end location"},
        "value": {"значение", "value"},
        "count": {"кол-во", "количество", "count"},
    }
    for idx, row in enumerate(rows[:30]):
        normalized = [_norm_header(cell) for cell in row]
        colmap: dict[str, int] = {}
        for col_idx, cell in enumerate(normalized):
            for target, names in aliases.items():
                if cell in names and target not in colmap:
                    colmap[target] = col_idx
        required = {"vehicle", "violation", "start_raw"}
        if required.issubset(colmap):
            return idx, colmap
    return 2, {
        "number": 0,
        "vehicle": 1,
        "rating": 2,
        "violation": 3,
        "start_raw": 4,
        "start_pos": 5,
        "end_raw": 6,
        "end_pos": 7,
        "value": 8,
        "count": 9,
    }


def _cell(row: tuple[Any, ...] | list[Any], colmap: dict[str, int], name: str, default: Any = None) -> Any:
    idx = colmap.get(name)
    if idx is None or idx >= len(row):
        return default
    return row[idx]


def _looks_like_event_row(vehicle: Any, violation: Any, start_raw: Any) -> bool:
    if not vehicle or not violation:
        return False
    vehicle_text = str(vehicle).strip().lower()
    violation_text = str(violation).strip().lower()
    if vehicle_text in {"итого", "тс", "группировка"}:
        return False
    if violation_text in {"нарушение", "-----"}:
        return False
    if start_raw in (None, ""):
        return True
    return parse_dt(start_raw) is not None or isinstance(start_raw, datetime)


def parse_excel_bytes(file_bytes: bytes, sheet_name: str | None = None) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    sheet_names = get_sheet_names(file_bytes)
    selected_sheet = pick_source_sheet_name(sheet_names, sheet_name)
    rows, links = _read_rows(file_bytes, selected_sheet)
    header_idx, colmap = _find_header_row(rows)
    vehicle_summary: list[dict[str, Any]] = []
    events: list[dict[str, Any]] = []

    for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
        if not any(cell is not None for cell in row):
            continue

        vehicle = _cell(row, colmap, "vehicle")
        violation = _cell(row, colmap, "violation")
        start_raw = _cell(row, colmap, "start_raw")
        end_raw = _cell(row, colmap, "end_raw")
        start_pos = _cell(row, colmap, "start_pos")
        end_pos = _cell(row, colmap, "end_pos")
        start_map_url = _cell_link(links, row_idx, colmap, "start_pos")
        end_map_url = _cell_link(links, row_idx, colmap, "end_pos")
        start_lat, start_lon = _coords_from_map_url(start_map_url)
        end_lat, end_lon = _coords_from_map_url(end_map_url)
        value = _cell(row, colmap, "value")
        count_raw = _cell(row, colmap, "count")
        rating_raw = _cell(row, colmap, "rating")

        if not vehicle or str(vehicle).strip().lower() in {"итого", "тс", "группировка"}:
            continue

        plate, route = split_vehicle(vehicle)

        if violation in (None, "", "-----"):
            count = safe_int(count_raw)
            rating = rating_raw if isinstance(rating_raw, (int, float)) else None
            vehicle_summary.append({
                "vehicle": str(vehicle),
                "plate": plate,
                "route": route,
                "rating": rating,
                "summary_count": count,
                "start_raw": start_raw,
                "end_raw": end_raw,
            })
            continue

        if not _looks_like_event_row(vehicle, violation, start_raw):
            continue

        dt = parse_dt(start_raw)
        severity = severity_class(violation)
        events.append({
            "vehicle": str(vehicle),
            "plate": plate,
            "route": route,
            "violation": str(violation),
            "action": action_type(violation),
            "severity": severity,
            "risk": risk_points(severity),
            "dt": dt,
            "start_raw": start_raw,
            "date": dt.date() if dt else pd.NaT,
            "hour": dt.hour if dt else None,
            "start_pos": start_pos,
            "start_lat": start_lat,
            "start_lon": start_lon,
            "start_map_url": start_map_url,
            "end_raw": end_raw,
            "end_pos": end_pos,
            "end_lat": end_lat,
            "end_lon": end_lon,
            "end_map_url": end_map_url,
            "value": value,
            "value_num": parse_num_value(value),
            "count": safe_int(count_raw, 1) or 1,
        })

    events_df = pd.DataFrame(events)
    if events_df.empty:
        events_df = pd.DataFrame(columns=EVENT_COLUMNS)

    events_df = normalize_events_df(events_df)
    summary_df = pd.DataFrame(vehicle_summary)
    vehicles_df = build_vehicle_summary(events_df, summary_df)
    return events_df, vehicles_df, selected_sheet


def normalize_events_df(events_df: pd.DataFrame) -> pd.DataFrame:
    df = events_df.copy()
    for col in EVENT_COLUMNS:
        if col not in df.columns:
            df[col] = None

    df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["hour"] = pd.to_numeric(df["hour"], errors="coerce")
    df["risk"] = pd.to_numeric(df["risk"], errors="coerce").fillna(0).astype(int)
    df["count"] = pd.to_numeric(df["count"], errors="coerce").fillna(1).astype(int)
    df["value_num"] = pd.to_numeric(df["value_num"], errors="coerce")
    for col in ["start_lat", "start_lon", "end_lat", "end_lon"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    text_cols = ["vehicle", "plate", "route", "violation", "action", "severity", "start_pos", "start_map_url", "end_pos", "end_map_url", "value"]
    for col in text_cols:
        df[col] = df[col].fillna("").astype(str)

    return df


# =========================
# SQLite storage + duplicate protection
# =========================

def ensure_parent_dir(path: str | Path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)


def init_db(db_path: str | Path) -> None:
    ensure_parent_dir(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("""
        CREATE TABLE IF NOT EXISTS events (
            event_hash TEXT PRIMARY KEY,
            vehicle TEXT,
            plate TEXT,
            route TEXT,
            violation TEXT,
            action TEXT,
            severity TEXT,
            risk INTEGER,
            dt TEXT,
            start_raw TEXT,
            date TEXT,
            hour INTEGER,
            start_pos TEXT,
            start_lat REAL,
            start_lon REAL,
            start_map_url TEXT,
            end_raw TEXT,
            end_pos TEXT,
            end_lat REAL,
            end_lon REAL,
            end_map_url TEXT,
            value TEXT,
            value_num REAL,
            count INTEGER,
            source_file TEXT,
            source_sheet TEXT,
            file_hash TEXT,
            imported_at TEXT
        )
        """)
        existing_cols = {row[1] for row in conn.execute("PRAGMA table_info(events)").fetchall()}
        migrations = {
            "start_lat": "REAL",
            "start_lon": "REAL",
            "start_map_url": "TEXT",
            "end_lat": "REAL",
            "end_lon": "REAL",
            "end_map_url": "TEXT",
        }
        for col, col_type in migrations.items():
            if col not in existing_cols:
                conn.execute(f"ALTER TABLE events ADD COLUMN {col} {col_type}")

        conn.execute("""
        CREATE TABLE IF NOT EXISTS imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT,
            source_sheet TEXT,
            file_hash TEXT,
            imported_at TEXT,
            parsed_events INTEGER,
            inserted_events INTEGER,
            duplicate_events INTEGER,
            min_dt TEXT,
            max_dt TEXT
        )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_events_dt ON events(dt)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_events_plate ON events(plate)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_events_violation ON events(violation)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_events_severity ON events(severity)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_events_geo_missing ON events(start_lat, start_lon)")
        conn.commit()


def file_sha256(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def _norm_for_hash(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        return value.isoformat()
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def make_event_hash(row: pd.Series | dict[str, Any]) -> str:
    # В Wialon-выгрузке нет стабильного ID события, поэтому используем натуральный ключ.
    # Он переживает повторные импорты одного и того же файла и пересекающиеся отчётные периоды.
    fields = [
        "plate",
        "violation",
        "dt",
        "start_pos",
        "end_pos",
        "value",
    ]
    raw = "|".join(_norm_for_hash(row.get(field, "")) for field in fields)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _to_db_text(value: Any) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, pd.Timestamp):
        return value.isoformat(sep=" ")
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value)


def _to_db_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def _to_db_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
        return int(value)
    except Exception:
        return None


def import_events_to_db(
    db_path: str | Path,
    events_df: pd.DataFrame,
    source_file: str,
    source_sheet: str,
    file_bytes: bytes | None = None,
) -> dict[str, Any]:
    init_db(db_path)
    df = normalize_events_df(events_df)
    parsed = int(len(df))
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    fhash = file_sha256(file_bytes) if file_bytes is not None else ""

    inserted = 0
    rows_for_insert: list[tuple[Any, ...]] = []

    for _, row in df.iterrows():
        event_hash = make_event_hash(row)
        rows_for_insert.append((
            event_hash,
            _to_db_text(row.get("vehicle")),
            _to_db_text(row.get("plate")),
            _to_db_text(row.get("route")),
            _to_db_text(row.get("violation")),
            _to_db_text(row.get("action")),
            _to_db_text(row.get("severity")),
            _to_db_int(row.get("risk")),
            _to_db_text(row.get("dt")),
            _to_db_text(row.get("start_raw")),
            _to_db_text(row.get("date")),
            _to_db_int(row.get("hour")),
            _to_db_text(row.get("start_pos")),
            _to_db_float(row.get("start_lat")),
            _to_db_float(row.get("start_lon")),
            _to_db_text(row.get("start_map_url")),
            _to_db_text(row.get("end_raw")),
            _to_db_text(row.get("end_pos")),
            _to_db_float(row.get("end_lat")),
            _to_db_float(row.get("end_lon")),
            _to_db_text(row.get("end_map_url")),
            _to_db_text(row.get("value")),
            _to_db_float(row.get("value_num")),
            _to_db_int(row.get("count")) or 1,
            source_file,
            source_sheet,
            fhash,
            now,
        ))

    with sqlite3.connect(db_path) as conn:
        for item in rows_for_insert:
            cur = conn.execute("""
                INSERT OR IGNORE INTO events (
                    event_hash, vehicle, plate, route, violation, action, severity, risk, dt, start_raw,
                    date, hour, start_pos, start_lat, start_lon, start_map_url,
                    end_raw, end_pos, end_lat, end_lon, end_map_url, value, value_num, count,
                    source_file, source_sheet, file_hash, imported_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, item)
            if cur.rowcount:
                inserted += cur.rowcount
            else:
                # Existing duplicate event from old imports may not have geo fields.
                # Fill coordinates/map URLs without changing the duplicate key.
                conn.execute("""
                    UPDATE events
                    SET
                        start_lat = COALESCE(start_lat, ?),
                        start_lon = COALESCE(start_lon, ?),
                        start_map_url = COALESCE(NULLIF(start_map_url, ''), ?),
                        end_lat = COALESCE(end_lat, ?),
                        end_lon = COALESCE(end_lon, ?),
                        end_map_url = COALESCE(NULLIF(end_map_url, ''), ?)
                    WHERE event_hash = ?
                """, (item[13], item[14], item[15], item[18], item[19], item[20], item[0]))

        min_dt = None
        max_dt = None
        if not df.empty:
            min_val = pd.to_datetime(df["dt"], errors="coerce").min()
            max_val = pd.to_datetime(df["dt"], errors="coerce").max()
            min_dt = _to_db_text(min_val)
            max_dt = _to_db_text(max_val)

        duplicate = parsed - inserted
        conn.execute("""
            INSERT INTO imports (
                source_file, source_sheet, file_hash, imported_at, parsed_events,
                inserted_events, duplicate_events, min_dt, max_dt
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (source_file, source_sheet, fhash, now, parsed, inserted, duplicate, min_dt, max_dt))
        conn.commit()

    return {
        "parsed_events": parsed,
        "inserted_events": inserted,
        "duplicate_events": parsed - inserted,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "file_hash": fhash,
        "imported_at": now,
    }



# =========================
# Wialon geo enrichment
# =========================

def _geo_norm_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace(" ", "").replace("-", "").replace("_", "")
    return text


def _parse_local_dt_for_wialon(value: Any, tz_offset_hours: float = 3.0) -> datetime | None:
    if value in (None, ""):
        return None
    dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return None
    if isinstance(dt, pd.Timestamp):
        py_dt = dt.to_pydatetime()
    elif isinstance(dt, datetime):
        py_dt = dt
    else:
        return None
    if py_dt.tzinfo is None:
        py_dt = py_dt.replace(tzinfo=timezone(timedelta(hours=float(tz_offset_hours))))
    return py_dt


def _message_time(message: dict[str, Any]) -> int | None:
    for key in ("t", "time", "tm"):
        value = message.get(key)
        if value is not None:
            try:
                return int(float(value))
            except Exception:
                pass
    return None


def _message_coords(message: dict[str, Any]) -> tuple[float | None, float | None]:
    pos = message.get("pos") if isinstance(message, dict) else None
    if isinstance(pos, dict):
        # Wialon usually stores longitude in x and latitude in y.
        lon = pos.get("x")
        lat = pos.get("y")
        try:
            if lat is not None and lon is not None:
                lat_f = float(lat)
                lon_f = float(lon)
                if -90 <= lat_f <= 90 and -180 <= lon_f <= 180:
                    return lat_f, lon_f
        except Exception:
            pass
    for lat_key, lon_key in (("lat", "lon"), ("y", "x")):
        try:
            lat = message.get(lat_key)
            lon = message.get(lon_key)
            if lat is not None and lon is not None:
                lat_f = float(lat)
                lon_f = float(lon)
                if -90 <= lat_f <= 90 and -180 <= lon_f <= 180:
                    return lat_f, lon_f
        except Exception:
            pass
    return None, None


def build_unit_id_by_plate(units: list[Any]) -> dict[str, int]:
    """Map report plate values to Wialon unit IDs by unit names."""
    result: dict[str, int] = {}
    for unit in units:
        name = str(getattr(unit, "name", "") or "")
        object_id = int(getattr(unit, "object_id"))
        norm = _geo_norm_text(name)
        if norm:
            result[norm] = object_id
        first = re.split(r"[\s(/]", name.strip(), maxsplit=1)[0]
        first_norm = _geo_norm_text(first)
        if first_norm and first_norm not in result:
            result[first_norm] = object_id
    return result


def match_unit_id_for_plate(plate: str, unit_map: dict[str, int]) -> int | None:
    target = _geo_norm_text(plate)
    if not target:
        return None
    if target in unit_map:
        return unit_map[target]
    for name_norm, object_id in unit_map.items():
        if target and (name_norm.startswith(target) or target in name_norm):
            return object_id
    return None


def _nearest_coords_from_messages(messages: list[dict[str, Any]], target_ts: int) -> tuple[float | None, float | None]:
    best: tuple[int, float, float] | None = None
    for message in messages:
        lat, lon = _message_coords(message)
        if lat is None or lon is None:
            continue
        tm = _message_time(message)
        diff = abs(int(tm or target_ts) - int(target_ts))
        if best is None or diff < best[0]:
            best = (diff, lat, lon)
    if best is None:
        return None, None
    return best[1], best[2]


def enrich_missing_geo_for_period(
    db_path: str | Path,
    client: Any,
    *,
    date_from: Any | None = None,
    date_to: Any | None = None,
    tz_offset_hours: float = 3.0,
    window_sec: int = 180,
    limit: int = 500,
    load_count: int = 50,
) -> dict[str, Any]:
    """Fill start/end coordinates by querying Wialon messages around event times."""
    init_db(db_path)
    units = client.search_units(mask="*", limit=0)
    unit_map = build_unit_id_by_plate(units)

    params: list[Any] = []
    where = "WHERE (start_lat IS NULL OR start_lon IS NULL) AND dt IS NOT NULL AND dt <> ''"
    if date_from is not None:
        where += " AND dt >= ?"
        params.append(str(date_from))
    if date_to is not None:
        where += " AND dt <= ?"
        params.append(str(date_to))

    query = f"""
        SELECT event_hash, plate, dt, end_raw
        FROM events
        {where}
        ORDER BY dt DESC
        LIMIT ?
    """
    params.append(int(limit))

    checked = 0
    updated = 0
    no_unit = 0
    no_message = 0
    api_errors = 0
    cache: dict[tuple[int, int], tuple[float | None, float | None]] = {}

    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        for event_hash, plate, dt_raw, end_raw in rows:
            checked += 1
            unit_id = match_unit_id_for_plate(str(plate or ""), unit_map)
            if not unit_id:
                no_unit += 1
                continue

            start_dt = _parse_local_dt_for_wialon(dt_raw, tz_offset_hours)
            if not start_dt:
                no_message += 1
                continue
            start_ts = int(start_dt.timestamp())
            cache_key = (unit_id, start_ts)
            if cache_key in cache:
                start_lat, start_lon = cache[cache_key]
            else:
                try:
                    messages = client.load_messages_interval(
                        item_id=unit_id,
                        time_from=start_ts - int(window_sec),
                        time_to=start_ts + int(window_sec),
                        load_count=int(load_count),
                    )
                    start_lat, start_lon = _nearest_coords_from_messages(messages, start_ts)
                    cache[cache_key] = (start_lat, start_lon)
                except Exception:
                    api_errors += 1
                    continue

            if start_lat is None or start_lon is None:
                no_message += 1
                continue

            start_url = f"https://www.google.com/maps?q={start_lat:.7f},{start_lon:.7f}"
            end_lat = None
            end_lon = None
            end_url = None

            end_dt = _parse_local_dt_for_wialon(end_raw, tz_offset_hours)
            if end_dt:
                end_ts = int(end_dt.timestamp())
                end_key = (unit_id, end_ts)
                if end_key in cache:
                    end_lat, end_lon = cache[end_key]
                else:
                    try:
                        end_messages = client.load_messages_interval(
                            item_id=unit_id,
                            time_from=end_ts - int(window_sec),
                            time_to=end_ts + int(window_sec),
                            load_count=int(load_count),
                        )
                        end_lat, end_lon = _nearest_coords_from_messages(end_messages, end_ts)
                        cache[end_key] = (end_lat, end_lon)
                    except Exception:
                        end_lat, end_lon = None, None
                if end_lat is not None and end_lon is not None:
                    end_url = f"https://www.google.com/maps?q={end_lat:.7f},{end_lon:.7f}"

            cur = conn.execute("""
                UPDATE events
                SET
                    start_lat = COALESCE(start_lat, ?),
                    start_lon = COALESCE(start_lon, ?),
                    start_map_url = COALESCE(NULLIF(start_map_url, ''), ?),
                    end_lat = COALESCE(end_lat, ?),
                    end_lon = COALESCE(end_lon, ?),
                    end_map_url = COALESCE(NULLIF(end_map_url, ''), ?)
                WHERE event_hash = ?
            """, (start_lat, start_lon, start_url, end_lat, end_lon, end_url, event_hash))
            if cur.rowcount:
                updated += 1
        conn.commit()

    return {
        "checked": checked,
        "updated": updated,
        "no_unit": no_unit,
        "no_message": no_message,
        "api_errors": api_errors,
        "units_loaded": len(units),
        "limit": int(limit),
        "window_sec": int(window_sec),
    }


def load_events_from_db(db_path: str | Path) -> pd.DataFrame:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        df = pd.read_sql_query("SELECT * FROM events", conn)
    if df.empty:
        return pd.DataFrame(columns=DB_EVENT_COLUMNS)
    df = normalize_events_df(df)
    for col in ["event_hash", "source_file", "source_sheet", "file_hash", "imported_at"]:
        if col not in df.columns:
            df[col] = ""
    return df[DB_EVENT_COLUMNS]


def get_db_stats(db_path: str | Path) -> dict[str, Any]:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        row = conn.execute("""
            SELECT
                COUNT(*) AS events_count,
                COUNT(DISTINCT plate) AS vehicles_count,
                MIN(dt) AS min_dt,
                MAX(dt) AS max_dt
            FROM events
        """).fetchone()
        import_count = conn.execute("SELECT COUNT(*) FROM imports").fetchone()[0]
    return {
        "events_count": int(row[0] or 0),
        "vehicles_count": int(row[1] or 0),
        "min_dt": row[2],
        "max_dt": row[3],
        "import_count": int(import_count or 0),
    }


def get_import_history(db_path: str | Path, limit: int = 20) -> pd.DataFrame:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        return pd.read_sql_query("""
            SELECT
                imported_at,
                source_file,
                source_sheet,
                parsed_events,
                inserted_events,
                duplicate_events,
                min_dt,
                max_dt
            FROM imports
            ORDER BY id DESC
            LIMIT ?
        """, conn, params=(limit,))


def clear_database(db_path: str | Path) -> None:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute("DELETE FROM events")
        conn.execute("DELETE FROM imports")
        conn.commit()


# =========================
# Analytics
# =========================

def build_vehicle_summary(events_df: pd.DataFrame, summary_df: pd.DataFrame | None = None) -> pd.DataFrame:
    base_cols = ["vehicle", "plate", "route", "rating", "summary_count"]
    if summary_df is None or summary_df.empty:
        summary_df = pd.DataFrame(columns=base_cols)
    for col in base_cols:
        if col not in summary_df.columns:
            summary_df[col] = None

    if events_df.empty:
        out = summary_df[base_cols].copy()
        for col in ["events", "dangerous", "sharp", "other", "risk_score"]:
            out[col] = 0
        out["priority"] = "Норма"
        return out

    grouped = events_df.groupby(["vehicle", "plate", "route"], dropna=False).agg(
        events=("violation", "size"),
        dangerous=("severity", lambda s: int(s.isin(["Опасное", "Критическое"]).sum())),
        sharp=("severity", lambda s: int((s == "Резкое").sum())),
        other=("severity", lambda s: int((s == "Прочее").sum())),
        risk_score=("risk", "sum"),
    ).reset_index()

    if summary_df.empty:
        out = grouped.copy()
        out["rating"] = None
        out["summary_count"] = out["events"]
    else:
        out = summary_df[base_cols].drop_duplicates(subset=["vehicle"]).merge(
            grouped,
            on=["vehicle", "plate", "route"],
            how="outer",
        )

    for col in ["events", "dangerous", "sharp", "other", "risk_score"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).astype(int)
    out["summary_count"] = pd.to_numeric(out["summary_count"], errors="coerce").fillna(out["events"]).astype(int)
    out["priority"] = out["risk_score"].apply(lambda x: priority_zone(int(x)))
    out["display_count"] = out["events"].where(out["events"] > 0, out["summary_count"])
    return out


def filter_events(
    events_df: pd.DataFrame,
    date_start=None,
    date_end=None,
    severities: list[str] | None = None,
    violations: list[str] | None = None,
    vehicles: list[str] | None = None,
    actions: list[str] | None = None,
    location_text: str | None = None,
) -> pd.DataFrame:
    df = events_df.copy()
    if df.empty:
        return df

    def _normalize_date_value(value, use_last=False):
        if isinstance(value, (tuple, list)):
            if len(value) == 0:
                return None
            if len(value) == 1:
                return value[0]
            return value[-1] if use_last else value[0]
        return value

    date_start = _normalize_date_value(date_start, use_last=False)
    date_end = _normalize_date_value(date_end, use_last=True)

    if date_start is not None:
        start = pd.Timestamp(date_start)
        df = df[df["dt"] >= start]
    if date_end is not None:
        end = pd.Timestamp(date_end) + pd.Timedelta(days=1)
        df = df[df["dt"] < end]
    if severities:
        df = df[df["severity"].isin(severities)]
    if violations:
        df = df[df["violation"].isin(violations)]
    if vehicles:
        df = df[df["plate"].isin(vehicles) | df["vehicle"].isin(vehicles)]
    if actions:
        df = df[df["action"].isin(actions)]
    if location_text:
        text = location_text.strip().lower()
        if text:
            df = df[df["start_pos"].fillna("").astype(str).str.lower().str.contains(re.escape(text), na=False)]
    return df


def compute_metrics(events_df: pd.DataFrame) -> dict[str, Any]:
    df = events_df.copy()
    total_events = int(len(df))
    total_vehicles = int(df["plate"].nunique()) if not df.empty else 0
    dangerous = int(df["severity"].isin(["Опасное", "Критическое"]).sum()) if not df.empty else 0
    sharp = int((df["severity"] == "Резкое").sum()) if not df.empty else 0
    avg_per_vehicle = round(total_events / total_vehicles, 1) if total_vehicles else 0

    if df.empty:
        return {
            "total_events": 0,
            "total_vehicles": 0,
            "dangerous": 0,
            "sharp": 0,
            "danger_share": 0,
            "avg_per_vehicle": 0,
            "top_vehicle": "—",
            "top_vehicle_count": 0,
            "top_violation": "—",
            "top_violation_count": 0,
            "top_count": pd.DataFrame(),
            "top_risk": pd.DataFrame(),
            "violation_counts": pd.DataFrame(),
            "severity_counts": pd.DataFrame(),
            "hourly_counts": pd.DataFrame(),
            "location_counts": pd.DataFrame(),
        }

    vehicle_stats = df.groupby(["plate", "vehicle"], dropna=False).agg(
        violations=("violation", "size"),
        dangerous=("severity", lambda s: int(s.isin(["Опасное", "Критическое"]).sum())),
        sharp=("severity", lambda s: int((s == "Резкое").sum())),
        risk_score=("risk", "sum"),
    ).reset_index()
    vehicle_stats["priority"] = vehicle_stats["risk_score"].apply(lambda x: priority_zone(int(x)))

    top_count = vehicle_stats.sort_values(
        ["violations", "dangerous", "risk_score"], ascending=False
    ).head(10).reset_index(drop=True)
    top_risk = vehicle_stats.sort_values(
        ["risk_score", "dangerous", "violations"], ascending=False
    ).head(10).reset_index(drop=True)

    violation_counts = df["violation"].value_counts().rename_axis("violation").reset_index(name="count")
    violation_counts["share"] = violation_counts["count"] / total_events
    severity_counts = df["severity"].value_counts().rename_axis("severity").reset_index(name="count")
    hourly_counts = df.dropna(subset=["hour"]).assign(hour=lambda x: x["hour"].astype(int)).groupby("hour").size().reset_index(name="count")
    location_counts = df[df["start_pos"].notna()].groupby("start_pos").size().sort_values(ascending=False).head(20).rename_axis("location").reset_index(name="count")

    top_vehicle_row = top_count.iloc[0] if not top_count.empty else None
    top_violation_row = violation_counts.iloc[0] if not violation_counts.empty else None

    return {
        "total_events": total_events,
        "total_vehicles": total_vehicles,
        "dangerous": dangerous,
        "sharp": sharp,
        "danger_share": dangerous / total_events if total_events else 0,
        "avg_per_vehicle": avg_per_vehicle,
        "top_vehicle": top_vehicle_row["plate"] if top_vehicle_row is not None else "—",
        "top_vehicle_count": int(top_vehicle_row["violations"]) if top_vehicle_row is not None else 0,
        "top_violation": top_violation_row["violation"] if top_violation_row is not None else "—",
        "top_violation_count": int(top_violation_row["count"]) if top_violation_row is not None else 0,
        "top_count": top_count,
        "top_risk": top_risk,
        "violation_counts": violation_counts,
        "severity_counts": severity_counts,
        "hourly_counts": hourly_counts,
        "location_counts": location_counts,
    }


def prepare_dynamic(events_df: pd.DataFrame, violation_type: str = "Все", group_by: str = "День") -> pd.DataFrame:
    df = events_df.copy()
    if df.empty:
        return pd.DataFrame(columns=["period", "count"])
    if violation_type and violation_type != "Все":
        df = df[df["violation"] == violation_type]
    df = df.dropna(subset=["dt"])
    if df.empty:
        return pd.DataFrame(columns=["period", "count"])

    if group_by == "Час":
        df["period"] = df["dt"].dt.floor("h")
    elif group_by == "Неделя":
        df["period"] = df["dt"].dt.to_period("W").apply(lambda p: p.start_time)
    else:
        df["period"] = df["dt"].dt.floor("D")

    return df.groupby("period").size().reset_index(name="count").sort_values("period")


def to_excel_bytes(events_df: pd.DataFrame, metrics: dict[str, Any]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_events = events_df.copy()
        for col in ["dt", "date"]:
            if col in export_events.columns:
                export_events[col] = pd.to_datetime(export_events[col], errors="coerce")
        export_events.rename(columns={
            "vehicle": "ТС",
            "plate": "Госномер",
            "route": "Маршрут / привязка",
            "violation": "Нарушение",
            "action": "Тип",
            "severity": "Класс",
            "risk": "Риск-балл",
            "dt": "Дата/время",
            "hour": "Час",
            "start_pos": "Нач. положение",
            "end_pos": "Кон. положение",
            "value": "Значение",
            "value_num": "Значение число",
            "source_file": "Источник",
            "imported_at": "Дата импорта",
        }).to_excel(writer, sheet_name="События", index=False)
        for key, sheet in [("top_count", "Топ нарушителей"), ("top_risk", "Топ риск")]:
            df = metrics.get(key)
            if isinstance(df, pd.DataFrame) and not df.empty:
                df.to_excel(writer, sheet_name=sheet, index=False)
        viol = metrics.get("violation_counts")
        if isinstance(viol, pd.DataFrame) and not viol.empty:
            viol.to_excel(writer, sheet_name="Типы нарушений", index=False)
    output.seek(0)
    return output.getvalue()


def newest_xlsx_bytes(folder: str | Path = ".") -> tuple[bytes, str] | None:
    folder = Path(folder)
    files = [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")]
    if not files:
        return None
    file = max(files, key=lambda p: p.stat().st_mtime)
    return file.read_bytes(), file.name

# =========================
# Localized export override
# =========================

def to_excel_bytes(events_df: pd.DataFrame, metrics: dict[str, Any], lang: str = "ru") -> bytes:
    """Export filtered data. Backward-compatible with old calls; lang='tr' localizes sheet/column names."""
    from localization import col_label, label_priority, label_violation, localize_event_values, normalize_lang

    lang = normalize_lang(lang)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_events = localize_event_values(events_df.copy(), lang)
        for col in ["dt", "date"]:
            if col in export_events.columns:
                export_events[col] = pd.to_datetime(export_events[col], errors="coerce")
        export_events.rename(columns={
            "vehicle": col_label("vehicle", lang),
            "plate": col_label("plate", lang),
            "route": col_label("route", lang),
            "violation": col_label("violation", lang),
            "action": col_label("action", lang),
            "severity": col_label("severity", lang),
            "risk": col_label("risk", lang),
            "dt": col_label("dt", lang),
            "hour": col_label("hour", lang),
            "start_pos": col_label("start_pos", lang),
            "start_lat": "Нач. широта" if lang == "ru" else "Başl. enlem",
            "start_lon": "Нач. долгота" if lang == "ru" else "Başl. boylam",
            "start_map_url": "Карта начала" if lang == "ru" else "Başlangıç haritası",
            "end_pos": col_label("end_pos", lang),
            "end_lat": "Кон. широта" if lang == "ru" else "Bitiş enlem",
            "end_lon": "Кон. долгота" if lang == "ru" else "Bitiş boylam",
            "end_map_url": "Карта конца" if lang == "ru" else "Bitiş haritası",
            "value": col_label("value", lang),
            "value_num": col_label("value_num", lang),
            "source_file": col_label("source_file", lang),
            "imported_at": col_label("imported_at", lang),
        }).to_excel(writer, sheet_name="Olaylar" if lang == "tr" else "События", index=False)

        for key, sheet_ru, sheet_tr in [("top_count", "Топ нарушителей", "Top ihlalciler"), ("top_risk", "Топ риск", "Top risk")]:
            df = metrics.get(key)
            if isinstance(df, pd.DataFrame) and not df.empty:
                out = df.copy()
                if "priority" in out.columns:
                    out["priority"] = out["priority"].map(lambda x: label_priority(x, lang))
                out = out.rename(columns={
                    "plate": col_label("plate", lang),
                    "vehicle": col_label("vehicle", lang),
                    "violations": col_label("violations", lang),
                    "dangerous": col_label("dangerous", lang),
                    "sharp": col_label("sharp", lang),
                    "risk_score": col_label("risk_score", lang),
                    "priority": col_label("priority", lang),
                })
                out.to_excel(writer, sheet_name=sheet_tr if lang == "tr" else sheet_ru, index=False)

        viol = metrics.get("violation_counts")
        if isinstance(viol, pd.DataFrame) and not viol.empty:
            out = viol.copy()
            if "violation" in out.columns:
                out["violation"] = out["violation"].map(lambda x: label_violation(x, lang))
            out = out.rename(columns={
                "violation": col_label("violation", lang),
                "count": col_label("count", lang),
                "share": col_label("share", lang),
            })
            out.to_excel(writer, sheet_name="İhlal türleri" if lang == "tr" else "Типы нарушений", index=False)
    output.seek(0)
    return output.getvalue()
